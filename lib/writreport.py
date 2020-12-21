import os
import time
import openpyxl
from openpyxl import Workbook


def createbook():
    project_name = "AutoTest"
    curPath = os.path.abspath(os.path.dirname(__file__))
    #print("当前路径： %s" %curPath)
    ReportPath = curPath.replace('lib','Report')
    #print("报告路径： %s" %ReportPath)
    localtime=time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time()))
    #print("时间： %s" %localtime)

    os.chdir(ReportPath)
    book = Workbook()
    sheet1 = book.create_sheet('测试报告')

    book.remove(book.get_sheet_by_name('Sheet'))

    sheet1.cell(1,1,"ID")
    sheet1.cell(1,2,"用例名称")
    sheet1.cell(1,3,"模式")
    sheet1.cell(1,4,"URL")
    sheet1.cell(1,5,"执行结果")
    sheet1.cell(1,6,"Response")
    name = "测试报告%s.xlsx" %localtime


    rename = ReportPath +"\\"+ name
    book.save(name)
    return rename
def writebook(name,data):

    oldbook = openpyxl.load_workbook(name,data_only=True)


    table = oldbook.get_sheet_by_name("测试报告")
    print(oldbook.sheetnames)


    rown = table.max_row

    nrown = rown + 1
    table.cell(nrown,1,rown)
    table.cell(nrown,2,data['name'])
    table.cell(nrown,3,data['method'])
    table.cell(nrown,4,data['url'])
    table.cell(nrown,5,data['result'])
    table.cell(nrown,6,data['Response'])
    oldbook.save(name)

    re = "报告写入成功"
    return re
